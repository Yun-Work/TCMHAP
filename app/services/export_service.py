import io
import base64
import pandas as pd
from sqlalchemy import text
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from datetime import datetime
from zoneinfo import ZoneInfo
from app.db import engine

# SQL：展開 organs(JSON 陣列) → 月份 × 器官 次數
SQL = text("""
SELECT
  DATE_FORMAT(fa.created_at, '%Y-%m-01') AS month,
  jt.organ AS organ,
  COUNT(*) AS symptom_count
FROM face_analysis fa
JOIN JSON_TABLE(
  fa.organs, '$[*]' COLUMNS (organ VARCHAR(50) PATH '$')
) jt
WHERE (:user_id IS NULL OR fa.user_id = :user_id)
  AND (:start   IS NULL OR fa.created_at >= :start)
  AND (:end     IS NULL OR fa.created_at < DATE_ADD(:end, INTERVAL 1 DAY))
GROUP BY month, organ
ORDER BY month, organ;
""")

# 查使用者
USER_SQL = text("""
SELECT name, gender, birth_date
FROM users
WHERE user_id = :uid
LIMIT 1
""")

def _fetch_user_row(conn, user_id: int | None):
    """抓取姓名 / 性別 / 生日並轉換格式"""
    if user_id is None:
        return None
    row = conn.execute(USER_SQL, {"uid": user_id}).fetchone()
    if not row:
        return None
    d = dict(row._mapping)

    # 性別轉中文
    gender_map = {"male": "男", "female": "女"}
    gender_display = gender_map.get((d.get("gender") or "").lower(), "-")

    # 生日轉中文格式
    birth_val = d.get("birth_date")
    if birth_val:
        try:
            birth_display = birth_val.strftime("%Y年%m月%d日")
        except Exception:
            birth_display = str(birth_val)
    else:
        birth_display = "-"

    return {
        "name": d.get("name") or "-",
        "gender": gender_display,
        "birth_date": birth_display,
    }

def _make_excel_stream(df: pd.DataFrame, user_row=None) -> io.BytesIO:
    """把查詢結果轉成 Excel，加入大標題 + 基本資料 + 彩色表格"""
    bio = io.BytesIO()

    # 沒資料時也匯出檔案
    if df.empty:
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            ws_name = "History"
            pd.DataFrame([{"訊息": "查無資料"}]).to_excel(writer, index=False, sheet_name=ws_name, startrow=5)
            ws = writer.sheets[ws_name]

            # ===== 大標題 =====
            ws["A1"] = "健康歷史月份紀錄表"
            ws.merge_cells("A1:D1")
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

            # ===== 基本資料 =====
            if user_row:
                ws["A2"] = "姓名";  ws["B2"] = user_row.get("name")
                ws["A3"] = "性別";  ws["B3"] = user_row.get("gender")
                ws["A4"] = "生日";  ws["B4"] = user_row.get("birth_date")
                for r in (2, 3, 4):
                    ws[f"A{r}"].font = Font(bold=True)
        bio.seek(0)
        return bio

    # 樞紐表
    df["month"] = pd.to_datetime(df["month"]).dt.to_period("M")
    pivot = df.pivot_table(index="organ", columns="month", values="symptom_count", fill_value=0)
    full_months = pd.period_range(pivot.columns.min(), pivot.columns.max(), freq="M")
    pivot = pivot.reindex(columns=full_months, fill_value=0).sort_index()
    pivot.columns = pivot.columns.astype(str)

    # 版面
    TABLE_TOP_ROW = 6

    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        ws_name = "History"
        df_out = pivot.reset_index()
        df_out.to_excel(writer, sheet_name=ws_name, index=False, startrow=TABLE_TOP_ROW - 1)
        ws = writer.sheets[ws_name]

        # ===== 大標題 =====
        end_col_letter = get_column_letter(df_out.shape[1])
        ws["A1"] = "健康歷史月份紀錄表"
        ws.merge_cells(f"A1:{end_col_letter}1")
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # ===== 基本資料 =====
        if user_row:
            ws["A2"] = "姓名";  ws["B2"] = user_row["name"]
            ws["A3"] = "性別";  ws["B3"] = user_row["gender"]
            ws["A4"] = "生日";  ws["B4"] = user_row["birth_date"]
            for r in (2, 3, 4):
                ws[f"A{r}"].font = Font(bold=True)
                ws[f"A{r}"].alignment = Alignment(horizontal="right")
                ws[f"B{r}"].alignment = Alignment(horizontal="left")

        # ===== 彩色條件式 =====
        nrows, ncols = df_out.shape
        data_first_row = TABLE_TOP_ROW + 1
        data_last_row = data_first_row + (nrows - 1)
        start_col, end_col = 2, ncols
        data_range = f"{get_column_letter(start_col)}{data_first_row}:{get_column_letter(end_col)}{data_last_row}"

        green  = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        yellow = PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid")
        red    = PatternFill(start_color="E53935", end_color="E53935", fill_type="solid")
        ws.conditional_formatting.add(data_range, CellIsRule(operator='greaterThan', formula=['3'], fill=red))
        ws.conditional_formatting.add(data_range, CellIsRule(operator='between', formula=['1','3'], fill=yellow))
        ws.conditional_formatting.add(data_range, CellIsRule(operator='equal', formula=['0'], fill=green))

        # ===== 色碼說明 =====
        legend_col = end_col + 2
        lc, lc2 = get_column_letter(legend_col), get_column_letter(legend_col+1)
        ws[f"{lc}{data_first_row}"].fill = green
        ws[f"{lc2}{data_first_row}"] = "綠色(0次): 無症狀紀錄，狀態穩定"
        ws[f"{lc}{data_first_row+1}"].fill = yellow
        ws[f"{lc2}{data_first_row+1}"] = "黃色(1–3次): 輕度症狀出現，建議觀察追蹤"
        ws[f"{lc}{data_first_row+2}"].fill = red
        ws[f"{lc2}{data_first_row+2}"] = "紅色(≥4次): 症狀較頻繁，建議深入檢查或諮詢醫師"

        # ===== 匯出日期時間 =====
        ts = datetime.now(ZoneInfo("Asia/Taipei")).strftime("%Y-%m-%d %H:%M:%S")
        footer_row = data_last_row + 2
        footer_range = f"A{footer_row}:{get_column_letter(end_col)}{footer_row}"
        ws.merge_cells(footer_range)
        ws[f"A{footer_row}"] = f"匯出日期時間：{ts}"
        ws[f"A{footer_row}"].font = Font(bold=True)
        ws[f"A{footer_row}"].alignment = Alignment(horizontal="left", vertical="center")

    bio.seek(0)
    return bio

def build_symptom_history_excel(user_id=None, start=None, end=None) -> io.BytesIO:
    with engine.begin() as conn:
        df = pd.read_sql(SQL, conn, params={"user_id": user_id, "start": start, "end": end})
        user_row = _fetch_user_row(conn, user_id)
    return _make_excel_stream(df, user_row)

def build_symptom_history_excel_base64(user_id=None, start=None, end=None) -> str:
    with engine.begin() as conn:
        df = pd.read_sql(SQL, conn, params={"user_id": user_id, "start": start, "end": end})
        user_row = _fetch_user_row(conn, user_id)
    bio = _make_excel_stream(df, user_row)
    return base64.b64encode(bio.getvalue()).decode("utf-8")
