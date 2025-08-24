# export_excel.py
import pandas as pd
from sqlalchemy import text
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from app.db import engine

# SQL：展開 JSON 陣列 → 月份 × 器官 次數
SQL = text("""
SELECT
  DATE_FORMAT(fa.created_at, '%Y-%m-01') AS month,
  jt.organ AS organ,
  COUNT(*) AS symptom_count
FROM face_analysis fa
JOIN JSON_TABLE(
  fa.organs, '$[*]' COLUMNS (organ VARCHAR(50) PATH '$')
) jt
GROUP BY month, organ
ORDER BY month, organ;
""")

def main():
    # 1. 查詢資料
    with engine.begin() as conn:
        df = pd.read_sql(SQL, conn)

    if df.empty:
        print("查無資料，請檢查 face_analysis 內容。")
        return

    # 2. 轉矩陣（器官 × 月份），補齊完整月份
    df["month"] = pd.to_datetime(df["month"]).dt.to_period("M")
    pivot = df.pivot_table(index="organ", columns="month",
                           values="symptom_count", fill_value=0)
    full_months = pd.period_range(pivot.columns.min(), pivot.columns.max(), freq="M")
    pivot = pivot.reindex(columns=full_months, fill_value=0).sort_index()

    # 3. 匯出 Excel
    output = "symptom_history.xlsx"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pivot.to_excel(writer, sheet_name="History")
        ws = writer.sheets["History"]

        # 數字區範圍（B2 開始）
        nrows, ncols = pivot.shape
        start_row, start_col = 2, 2
        end_row = start_row + nrows - 1
        end_col = start_col + ncols - 1
        data_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

        # 顏色規則
        green  = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        yellow = PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid")
        red    = PatternFill(start_color="E53935", end_color="E53935", fill_type="solid")

        ws.conditional_formatting.add(data_range, CellIsRule(operator='equal', formula=['0'], fill=green))
        ws.conditional_formatting.add(data_range, CellIsRule(operator='between', formula=['1','3'], fill=yellow))
        ws.conditional_formatting.add(data_range, CellIsRule(operator='greaterThan', formula=['3'], fill=red))

        # 右側圖例
        legend_col = end_col + 2
        lc, lc2 = get_column_letter(legend_col), get_column_letter(legend_col + 1)
        ws[f"{lc}{start_row}"] = "色碼"
        ws[f"{lc2}{start_row}"] = "說明"
        ws[f"{lc}{start_row+1}"].fill = green
        ws[f"{lc2}{start_row+1}"] = "綠色 (=0)：無症狀紀錄，狀態穩定"
        ws[f"{lc}{start_row+2}"].fill = yellow
        ws[f"{lc2}{start_row+2}"] = "黃色 (1–3)：輕度症狀出現，建議觀察追蹤"
        ws[f"{lc}{start_row+3}"].fill = red
        ws[f"{lc2}{start_row+3}"] = "紅色 (≥4)：症狀較頻繁，建議深入檢查或諮詢醫師"

    print(f"已匯出 Excel：{output}")

if __name__ == "__main__":
    main()